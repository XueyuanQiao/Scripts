#!/usr/bin/env python3
# -*- coding: utf-8 -*-
'''
@author: Qiaoxueyuan
@time: 2017/11/1 11:25
'''

# import xlrd
# from xlutils.copy import copy
import openpyxl
import sys


def generate(max):
    # excel = xlrd.open_workbook("number.xls", formatting_info=True)
    # wexcel = copy(excel)
    # table = wexcel.get_sheet(0)
    num = len(str(max))
    excel = openpyxl.Workbook()
    table = excel.active
    num_of_8 = int(10 - int(num))
    behind_number = "1"
    while num_of_8:
        behind_number = behind_number + "8"
        num_of_8 = num_of_8 - 1
    if num >= 1:
        last = "0"
    i = num
    while i:
        last = last + "0"
        i = i - 1
    if len(behind_number) != 11:
        for add in range(0, int(max)):
            last_number = last[:len(last) - 1 - len(str(add))] + str(add)
            number = behind_number + last_number
            # table.write(add + 1, 1, str(number))
            table.cell(row=add + 2, column=1).value = str("客户" + str(add))
            table.cell(row=add + 2, column=2).value = str(number)
            table.cell(row=add + 2, column=3).value = str("这是第%d个描述:%s" % (add, number))
            # wexcel.save('number.xls')
    excel.save('number.xlsx')
    print("over,total = %d" % (add + 1))


if __name__ == "__main__":
    try:
        generate(int(sys.argv[1]))
    except IndexError:
        generate(100000)
    except:
        print("请输入正确数量参数！")
